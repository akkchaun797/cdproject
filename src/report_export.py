"""Build a downloadable, colour-coded Excel validation report."""
import io
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
AMBER = PatternFill("solid", fgColor="FFEB9C")
HEADER = PatternFill("solid", fgColor="1F4E78")
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
HEAD_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _cell_fill(value: str):
    v = str(value)
    if "✅" in v:
        return GREEN
    if "❌" in v:
        return RED
    if "⚠️" in v:
        return AMBER
    return None


def _write_sheet(ws, title: str, df: pd.DataFrame, start_row: int = 1) -> int:
    ws.cell(row=start_row, column=1, value=title).font = TITLE_FONT
    r = start_row + 2
    # header
    for c, col in enumerate(df.columns, 1):
        cell = ws.cell(row=r, column=c, value=col)
        cell.fill = HEADER
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    # body
    for _, row in df.iterrows():
        r += 1
        for c, col in enumerate(df.columns, 1):
            val = row[col]
            cell = ws.cell(row=r, column=c, value=str(val))
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            fill = _cell_fill(val)
            if fill:
                cell.fill = fill
    # column widths
    for c, col in enumerate(df.columns, 1):
        width = max(len(str(col)), *(len(str(v)) for v in df[col])) if len(df) else len(str(col))
        ws.column_dimensions[ws.cell(row=r, column=c).column_letter].width = min(max(width + 2, 14), 48)
    return r + 2


def build_excel_report(email_df: pd.DataFrame = None,
                       email_extras: dict = None,
                       social_df: pd.DataFrame = None) -> bytes:
    """Returns xlsx bytes with separate Email and Social Media report sheets."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M")

    if email_df is not None:
        ws = wb.create_sheet("Email PDF Report")
        nxt = _write_sheet(ws, f"BankABC — Email Creative Validation  ({stamp})", email_df)
        if email_extras:
            ws.cell(row=nxt, column=1, value="CTA / branding check:").font = Font(bold=True)
            ws.cell(row=nxt, column=2, value=str(email_extras.get("cta", "")))
            miss = email_extras.get("missing", [])
            ws.cell(row=nxt + 1, column=1, value="Merchants in Excel not in creative:").font = Font(bold=True)
            ws.cell(row=nxt + 1, column=2,
                    value=", ".join(miss) if miss else "None — all Excel merchants present ✅")

    if social_df is not None:
        ws = wb.create_sheet("Social Media Report")
        _write_sheet(ws, f"BankABC — Social Media Validation  ({stamp})", social_df)

    if not wb.sheetnames:
        wb.create_sheet("Report").cell(row=1, column=1, value="No data validated.")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
